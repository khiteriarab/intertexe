#!/usr/bin/env python3
"""Fill Contacted / HQ status on Contact List.xlsx from live hq_contacts. Does not invent emails."""
from __future__ import annotations

import json
import os
import re
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)
XLSX = "/Users/khiteri/Desktop/Contact List.xlsx"
HEADER_FILL = PatternFill("solid", fgColor="111111")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NEW_FILL = PatternFill("solid", fgColor="FFF3CD")


def load_env():
    env = {}
    root = Path(__file__).resolve().parents[1]
    for f in [root / ".env.vercel.local", root / ".env.local", root / ".env.development.local"]:
        if not f.exists():
            continue
        for line in f.read_text().splitlines():
            if not line.strip() or line.strip().startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env.setdefault(k.strip(), v.strip().strip('"').strip("'"))
    url = env.get("SUPABASE_URL") or env.get("NEXT_PUBLIC_SUPABASE_URL") or env.get("VITE_SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise SystemExit("Need SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY")
    return url.rstrip("/"), key


def rest(url, key, method, path, payload=None, params=None, extra_headers=None):
    q = "&".join(f"{k}={urllib.parse.quote(str(v), safe='')}" for k, v in (params or {}).items())
    full = f"{url}/rest/v1/{path}" + (f"?{q}" if q else "")
    data = None if payload is None else json.dumps(payload).encode()
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    if extra_headers:
        headers.update(extra_headers)
    req = urllib.request.Request(full, data=data, method=method, headers=headers)
    with urllib.request.urlopen(req, timeout=60) as res:
        body = res.read().decode()
        return json.loads(body) if body else None, dict(res.headers)


def rest_all(url, key, path, params):
    rows = []
    start = 0
    page = 1000
    while True:
        data, _headers = rest(
            url,
            key,
            "GET",
            path,
            params=params,
            extra_headers={"Range": f"{start}-{start + page - 1}", "Prefer": "count=exact"},
        )
        chunk = data or []
        rows.extend(chunk)
        if len(chunk) < page:
            break
        start += page
    return rows


def emails_from(value):
    cleaned = re.sub(r"@{2,}", "@", str(value or ""))
    return [e.lower() for e in EMAIL_RE.findall(cleaned)]


def fmt_contacted(hit):
    if not hit:
        return "Not in HQ", ""
    status = hit.get("outreach_status") or "not_contacted"
    last = hit.get("last_contacted_at")
    day = last[:10] if last else ""
    if status == "converted":
        return ("Yes — converted" + (f" ({day})" if day else "")), status
    if last or status not in ("not_contacted", None, ""):
        label = "Yes" if last else "No"
        if day:
            label = f"Yes — {day}"
        if status and status not in ("contacted", "not_contacted"):
            label = f"{label} ({status})"
        return label, status
    return "No", status


def lookup(by_email, raw):
    found = emails_from(raw)
    hits = [by_email[e] for e in found if e in by_email]
    if not hits:
        return None, found
    hits.sort(key=lambda h: h.get("last_contacted_at") or "")
    return hits[-1], found


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True)


def main():
    url, key = load_env()
    ws_rows = rest(url, key, "GET", "hq_workspaces", params={"slug": "eq.intertexe", "select": "id"})[0]
    workspace_id = ws_rows[0]["id"]
    contacts = rest_all(
        url,
        key,
        "hq_contacts",
        {
            "workspace_id": f"eq.{workspace_id}",
            "select": "normalized_email,contact_type,outreach_status,last_contacted_at,first_name,last_name,full_name,company_name",
        },
    )
    by_email = {c["normalized_email"]: c for c in contacts if c.get("normalized_email")}

    try:
        rest(
            url,
            key,
            "POST",
            "hq_outreach_statuses",
            payload={"key": "undeliverable", "label": "Undeliverable", "sort_order": 75},
            extra_headers={"Prefer": "resolution=merge-duplicates,return=minimal"},
        )
    except Exception:
        pass

    wb = load_workbook(XLSX)
    new_rows = []
    stats = {}

    # Brands
    brands = wb["Brands"]
    brands.cell(1, 8, "Contacted")
    brands.cell(1, 9, "HQ status")
    style_header(brands.cell(1, 8))
    style_header(brands.cell(1, 9))
    b_new = b_yes = b_no = 0
    for r in range(2, brands.max_row + 1):
        if not any(brands.cell(r, c).value not in (None, "") for c in range(1, 8)):
            continue
        hit, found = lookup(by_email, brands.cell(r, 4).value)
        label, status = fmt_contacted(hit)
        brands.cell(r, 8, label)
        brands.cell(r, 9, status)
        if label == "Not in HQ" and found:
            b_new += 1
            new_rows.append(("Brands", brands.cell(r, 1).value, found[0], "email not in HQ"))
            brands.cell(r, 8).fill = NEW_FILL
        elif label.startswith("Yes"):
            b_yes += 1
        else:
            b_no += 1
    stats["Brands"] = {"new": b_new, "contacted": b_yes, "not_contacted": b_no}

    # Organization
    org = wb["Organization"]
    org.cell(1, 7, "In HQ")
    org.cell(1, 8, "Contacted")
    org.cell(1, 9, "HQ status")
    org.cell(1, 10, "Note")
    for c in range(7, 11):
        style_header(org.cell(1, c))
    o_new = o_need = 0
    for r in range(2, org.max_row + 1):
        vals = [org.cell(r, c).value for c in range(1, 7)]
        if not any(v not in (None, "") for v in vals):
            continue
        email_raw = org.cell(r, 5).value
        if isinstance(email_raw, str) and "@@" in email_raw:
            org.cell(r, 5, re.sub(r"@{2,}", "@", email_raw).strip())
            org.cell(r, 10, "Fixed double @@ in email")
        hit, found = lookup(by_email, org.cell(r, 5).value)
        name = org.cell(r, 1).value
        first = org.cell(r, 2).value
        if found:
            label, status = fmt_contacted(hit)
            org.cell(r, 7, "Yes" if hit else "No — new")
            org.cell(r, 8, label)
            org.cell(r, 9, status)
            if not hit:
                o_new += 1
                new_rows.append(("Organization", f"{name} / {first}", found[0], "new person with email"))
                for c in range(1, 11):
                    if org.cell(r, c).value not in (None, ""):
                        org.cell(r, c).fill = NEW_FILL
        else:
            o_need += 1
            org.cell(r, 7, "No")
            org.cell(r, 8, "")
            org.cell(r, 9, "")
            note = org.cell(r, 10).value
            org.cell(r, 10, note or "Needs email — not imported")
    stats["Organization"] = {"new_with_email": o_new, "needs_email": o_need}

    # Influencers — header is row 3
    inf = wb["Influencers"]
    inf.cell(3, 6, "Contacted")
    inf.cell(3, 10, "HQ status")
    style_header(inf.cell(3, 6))
    style_header(inf.cell(3, 10))
    i_new = i_yes = i_no = 0
    for r in range(4, inf.max_row + 1):
        if not any(inf.cell(r, c).value not in (None, "") for c in range(1, 10)):
            continue
        hit, found = lookup(by_email, inf.cell(r, 5).value)
        label, status = fmt_contacted(hit)
        inf.cell(r, 6, label)
        inf.cell(r, 10, status)
        if label == "Not in HQ" and found:
            i_new += 1
            new_rows.append(("Influencers", f"{inf.cell(r,1).value} {inf.cell(r,2).value}", found[0], "email not in HQ"))
            inf.cell(r, 6).fill = NEW_FILL
        elif str(label).startswith("Yes"):
            i_yes += 1
        else:
            i_no += 1
    stats["Influencers"] = {"new": i_new, "contacted": i_yes, "not_contacted": i_no}

    # Customers
    cust = wb["Customers"]
    cust.cell(1, 4, "Contacted")
    cust.cell(1, 5, "HQ status")
    style_header(cust.cell(1, 4))
    style_header(cust.cell(1, 5))
    c_new = c_yes = c_no = c_missing_last = 0
    for r in range(2, cust.max_row + 1):
        if not any(cust.cell(r, c).value not in (None, "") for c in range(1, 4)):
            continue
        if not cust.cell(r, 2).value:
            c_missing_last += 1
        hit, found = lookup(by_email, cust.cell(r, 3).value)
        label, status = fmt_contacted(hit)
        cust.cell(r, 4, label)
        cust.cell(r, 5, status)
        if label == "Not in HQ" and found:
            c_new += 1
            new_rows.append(("Customers", cust.cell(r, 1).value, found[0], "email not in HQ"))
            cust.cell(r, 4).fill = NEW_FILL
        elif str(label).startswith("Yes"):
            c_yes += 1
        else:
            c_no += 1
    stats["Customers"] = {
        "new": c_new,
        "contacted": c_yes,
        "not_contacted": c_no,
        "missing_last_name": c_missing_last,
    }

    if "What's new" in wb.sheetnames:
        del wb["What's new"]
    overview = wb.create_sheet("What's new", 0)
    overview["A1"] = "What's new vs HQ (filled " + datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC") + ")"
    overview["A1"].font = Font(bold=True, size=14)
    overview.merge_cells("A1:D1")
    lines = [
        ("A3", "What changed"),
        (
            "A4",
            "Organization tab now has people + emails (Aug 14 import had 0 organization emails). Those 8 are highlighted yellow until they are in hq_contacts.",
        ),
        ("A5", "Influencers Contacted was blank; it is now filled from HQ last_contacted_at / outreach_status."),
        ("A6", "Customers and Brands got Contacted + HQ status the same way. Yellow = email not in HQ yet."),
        ("A7", "CFDA email had a typo (membership@@cfda.com) — corrected to membership@cfda.com. No other emails were invented."),
        ("A8", "Organization rows without an email were left blank and marked Needs email — not imported."),
        ("A9", "Customer last names that were blank were left blank."),
        ("A11", "New or missing-from-HQ emails"),
        ("A12", "Tab"),
        ("B12", "Name"),
        ("C12", "Email"),
        ("D12", "Why"),
    ]
    for cell, val in lines:
        overview[cell] = val
    overview["A3"].font = Font(bold=True)
    overview["A11"].font = Font(bold=True)
    for col in range(1, 5):
        style_header(overview.cell(12, col))
    for i, row in enumerate(new_rows, start=13):
        overview.cell(i, 1, row[0])
        overview.cell(i, 2, row[1])
        overview.cell(i, 3, row[2])
        overview.cell(i, 4, row[3])
        for c in range(1, 5):
            overview.cell(i, c).fill = NEW_FILL
    start = 13 + max(len(new_rows), 1) + 2
    overview.cell(start, 1, "Counts")
    overview.cell(start, 1).font = Font(bold=True)
    overview.cell(start + 1, 1, json.dumps(stats, indent=2))
    overview.cell(start + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")
    overview.row_dimensions[start + 1].height = 90
    for col, width in enumerate([22, 44, 42, 36], start=1):
        overview.column_dimensions[get_column_letter(col)].width = width

    wb.save(XLSX)
    print(json.dumps({"saved": XLSX, "hq_contacts": len(by_email), "new_or_missing": new_rows, "stats": stats}, indent=2))


if __name__ == "__main__":
    main()
