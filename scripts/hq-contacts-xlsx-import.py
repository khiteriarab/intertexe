#!/usr/bin/env python3
"""One-time Contact List.xlsx → hq_contacts. Never emails. Preview by default; --apply writes."""
from __future__ import annotations

import json
import os
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)
APPLY = "--apply" in sys.argv
XLSX = next((a for a in sys.argv[1:] if a != "--apply"), "/Users/khiteri/Desktop/Contact List.xlsx")

TAB_TYPE = {
    "customers": "customer",
    "influencers": "influencer",
    "businesses": "business",
    "brands": "brand",
    "partners": "organization",
}


def load_env():
    env = {}
    root = Path(__file__).resolve().parents[1]
    for f in [root / ".env.vercel.local", root / ".env.local", root / ".env.production"]:
        if not f.exists():
            continue
        for line in f.read_text().splitlines():
            if not line.strip() or line.strip().startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env.setdefault(k.strip(), v.strip().strip('"').strip("'"))
    url = env.get("SUPABASE_URL") or env.get("NEXT_PUBLIC_SUPABASE_URL") or env.get("VITE_SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise SystemExit("Need SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY")
    return url.rstrip("/"), key


def rest(url, key, method, path, payload=None, params=None):
    q = "&".join(f"{k}={urllib.parse.quote(str(v))}" for k, v in (params or {}).items())
    full = f"{url}/rest/v1/{path}" + (f"?{q}" if q else "")
    data = None if payload is None else json.dumps(payload).encode()
    req = urllib.request.Request(
        full,
        data=data,
        method=method,
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        },
    )
    with urllib.request.urlopen(req, timeout=60) as res:
        body = res.read().decode()
        return json.loads(body) if body else None


import urllib.parse  # noqa: E402


def blank(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def emails_from(value):
    return sorted({e.lower() for e in EMAIL_RE.findall(str(value or ""))})


def header_map(row):
    m = {}
    for i, c in enumerate(row or []):
        key = str(c or "").strip().lower().replace(" ", "_")
        if key:
            m[key] = i
    return m


def col(m, row, names):
    for n in names:
        i = m.get(n)
        if i is not None and i < len(row) and row[i] not in (None, ""):
            return blank(row[i])
    return None


def type_from_tab(name):
    return TAB_TYPE.get(name.strip().lower(), "other")


def parse_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    parsed = []
    invalid = []
    tab_counts = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header_idx = 0
        for i, r in enumerate(rows[:8]):
            joined = " ".join(str(c or "").lower() for c in r)
            if "email" in joined:
                header_idx = i
                break
        hmap = header_map(rows[header_idx])
        if "email" not in hmap and not any("mail" in k for k in hmap):
            empty_n = sum(1 for r in rows if any(c not in (None, "") for c in r))
            invalid.append({"tab": sheet_name, "reason": "no_email_column", "rows": empty_n})
            tab_counts[sheet_name] = {"rows": empty_n, "valid": 0}
            continue
        contact_type = type_from_tab(sheet_name)
        valid = 0
        for r in rows[header_idx + 1 :]:
            if not any(c not in (None, "") for c in r):
                continue
            found = emails_from(col(hmap, r, ["email", "e-mail", "email_address", "mail"]) or "")
            if not found:
                # some brand sheets put email only in the email col; try whole row
                found = emails_from(" ".join(str(c or "") for c in r))
            if not found:
                invalid.append({"tab": sheet_name, "reason": "missing_email"})
                continue
            first = col(hmap, r, ["first_name", "firstname", "first"])
            last = col(hmap, r, ["last_name", "lastname", "last"])
            full = col(hmap, r, ["full_name", "fullname", "name"]) or " ".join(x for x in [first, last] if x) or None
            company = col(hmap, r, ["brand", "company", "company_name", "organization"])
            for email in found:
                parsed.append(
                    {
                        "email": email,
                        "normalized_email": email,
                        "first_name": first,
                        "last_name": last,
                        "full_name": full,
                        "name": full,
                        "company_name": company,
                        "website": col(hmap, r, ["website", "url", "site"]),
                        "phone": col(hmap, r, ["phone", "tel", "telephone"]),
                        "instagram": col(hmap, r, ["instagram", "ig", "insta"]),
                        "tiktok": col(hmap, r, ["tiktok", "tt"]),
                        "city": col(hmap, r, ["city"]),
                        "country": col(hmap, r, ["country"]),
                        "notes": col(hmap, r, ["notes", "note", "reference", "estimated_tier"]),
                        "contact_type": contact_type,
                        "sheet_tab": sheet_name,
                        "source": "xlsx_import",
                        "marketing_eligible": False,
                    }
                )
                valid += 1
        tab_counts[sheet_name] = {"rows": sum(1 for r in rows[header_idx + 1 :] if any(c not in (None, "") for c in r)), "valid": valid}
    wb.close()
    return parsed, invalid, tab_counts


def merge_by_email(rows):
    by = {}
    ambiguous = []
    dupes = 0
    for row in rows:
        email = row["normalized_email"]
        prev = by.get(email)
        if not prev:
            by[email] = {**row, "tabs": [row["sheet_tab"]]}
            continue
        dupes += 1
        prev["tabs"].append(row["sheet_tab"])
        if prev["contact_type"] != row["contact_type"] and prev["contact_type"] != "other":
            types = sorted({prev["contact_type"], row["contact_type"]})
            ambiguous.append({"email": email, "types": types, "tabs": prev["tabs"]})
            extra = f"[AMBIGUOUS type: {' + '.join(types)} tabs={','.join(prev['tabs'])}]"
            prev["notes"] = " ".join(x for x in [prev.get("notes"), extra, row.get("notes")] if x)
            prev["contact_type"] = "other"
        for k in ["first_name", "last_name", "full_name", "company_name", "website", "phone", "instagram", "tiktok", "city", "country"]:
            if not prev.get(k) and row.get(k):
                prev[k] = row[k]
        prev["sheet_tab"] = ",".join(dict.fromkeys(prev["tabs"]))
    # fix ambiguous notes that used overwritten type
    return list(by.values()), dupes, ambiguous


def main():
    parsed, invalid, tab_counts = parse_xlsx(XLSX)
    unique, dupes, ambiguous = merge_by_email(parsed)
    url, key = load_env()
    ws = rest(url, key, "GET", "hq_workspaces", params={"slug": "eq.intertexe", "select": "id"})
    workspace_id = ws[0]["id"]
    existing = rest(
        url,
        key,
        "GET",
        "hq_contacts",
        params={"workspace_id": f"eq.{workspace_id}", "select": "id,normalized_email,contact_type,notes,outreach_status,user_id"},
    ) or []
    existing_by = {r["normalized_email"]: r for r in existing}
    prefs = rest(url, key, "GET", "user_preferences", params={"select": "user_id,email"}) or []
    users = {str(p.get("email") or "").strip().lower(): p["user_id"] for p in prefs if p.get("email") and p.get("user_id")}

    insert = []
    already = []
    already_account = []
    for row in unique:
        email = row["normalized_email"]
        hit = existing_by.get(email)
        uid = users.get(email) or (hit or {}).get("user_id")
        if hit:
            already.append({"email": email, "existingType": hit.get("contact_type"), "incomingType": row["contact_type"], "hasAccount": bool(uid)})
            continue
        rec = {**row, "workspace_id": workspace_id}
        rec.pop("tabs", None)
        if uid:
            rec["user_id"] = uid
            rec["outreach_status"] = "converted"
            already_account.append({"email": email, "contact_type": row["contact_type"]})
        else:
            rec["outreach_status"] = "not_contacted"
            rec["user_id"] = None
        insert.append(rec)

    type_counts = {}
    for r in insert:
        type_counts[r["contact_type"]] = type_counts.get(r["contact_type"], 0) + 1

    preview = {
        "apply": APPLY,
        "file": XLSX,
        "tabs": tab_counts,
        "counts": {
            "total_parsed_emails": len(parsed),
            "unique_valid_emails": len(unique),
            "duplicates_in_file": dupes,
            "invalid": len(invalid),
            "existing_hq_contacts": len(already),
            "existing_intertexe_accounts_on_insert": len(already_account),
            "insert": len(insert),
            "ambiguous_type_conflicts": len(ambiguous),
            **{f"insert_{k}": v for k, v in sorted(type_counts.items())},
        },
        "ambiguous": ambiguous[:20],
        "sample_insert": [
            {k: r[k] for k in ["email", "first_name", "last_name", "company_name", "contact_type", "sheet_tab", "city", "instagram", "tiktok", "website", "outreach_status"]}
            for r in insert[:8]
        ],
        "already_in_supabase": already[:8],
        "partners_note": "Partners tab has no email column and is not imported.",
    }
    print(json.dumps(preview, indent=2))
    if not APPLY:
        print("\nPreview only. Re-run with --apply to write. Will not email anyone.")
        return

    upserted = 0
    errors = []
    for row in insert:
        try:
            data = rest(url, key, "POST", "hq_contacts", payload=row)
            cid = (data or [{}])[0].get("id")
            if not cid:
                errors.append(row["email"])
                continue
            upserted += 1
            rest(
                url,
                key,
                "POST",
                "hq_contact_outreach",
                payload={
                    "contact_id": cid,
                    "email": row["email"],
                    "channel": "system",
                    "direction": "system",
                    "provider": "xlsx_import",
                    "event_type": "contact_imported",
                    "metadata": {"sheet_tab": row.get("sheet_tab")},
                },
            )
        except urllib.error.HTTPError as e:
            errors.append(f"{row['email']}: {e.read().decode()[:180]}")
    # backfill account links
    # RPC
    req = urllib.request.Request(
        f"{url}/rest/v1/rpc/hq_link_existing_users_to_contacts",
        data=json.dumps({"p_workspace_id": workspace_id}).encode(),
        method="POST",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
        },
    )
    linked = None
    try:
        with urllib.request.urlopen(req, timeout=60) as res:
            linked = json.loads(res.read().decode() or "null")
    except Exception as e:
        errors.append(f"link_rpc: {e}")
    print(json.dumps({"applied": True, "attempted": len(insert), "upserted": upserted, "linked_existing_users": linked, "errors": errors[:12]}, indent=2))


if __name__ == "__main__":
    main()
